import json
from datetime import timedelta, time, datetime, date
from decimal import Decimal

from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, F
from django.db.models.functions import TruncDate
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from django.views.decorators.http import require_POST

import openpyxl
from openpyxl.utils import get_column_letter

from .models import Product, Category, Order, OrderItem, Operator, CashWithdrawal, Printer


# --- Utility: serata (giorno scatta alle 05:00) ---
NIGHT_CUTOFF = time(5, 0)


def get_session_date(dt=None):
    """Restituisce la data di 'serata': se prima delle 5:00, conta come giorno precedente."""
    if dt is None:
        dt = timezone.localtime(timezone.now())
    if dt.time() < NIGHT_CUTOFF:
        return (dt - timedelta(days=1)).date()
    return dt.date()


def session_range(session_date):
    """Restituisce (start, end) datetime per una serata."""
    tz = timezone.get_current_timezone()
    start = timezone.make_aware(datetime.combine(session_date, NIGHT_CUTOFF), tz)
    end = start + timedelta(days=1)
    return start, end


@login_required
def cassa(request):
    products = Product.objects.select_related("category").filter(is_shortcut=True).order_by("category__name", "name")
    categories = Category.objects.all()
    operators = Operator.objects.order_by("name")
    return render(request, "cassa/cassa.html", {"products": products, "categories": categories, "operators": operators})


@login_required
def search_products(request):
    q = request.GET.get("q", "")
    products = Product.objects.filter(name__icontains=q).values("id", "name", "price", "icon", "stock", "print_destinations", "category__name")[:20]
    return JsonResponse(list(products), safe=False)


@login_required
@require_POST
def create_order(request):
    data = json.loads(request.body)
    items = data.get("items", [])
    payment = data.get("payment_method", "cash")
    discount = Decimal(str(data.get("discount_percent", 0)))

    if not items:
        return JsonResponse({"error": "Nessun prodotto"}, status=400)

    # Controllo stock
    for item in items:
        product = get_object_or_404(Product, pk=item["id"])
        qty = int(item.get("quantity", 1))
        if product.stock is not None and product.stock < qty:
            return JsonResponse({"error": f"{product.name} esaurito (disponibili: {product.stock})"}, status=400)

    order = Order.objects.create(user=request.user, payment_method=payment, discount_percent=discount, note=data.get("note", ""))

    subtotal = Decimal("0")
    for item in items:
        product = get_object_or_404(Product, pk=item["id"])
        qty = int(item.get("quantity", 1))
        OrderItem.objects.create(
            order=order, product=product, product_name=product.name, price=product.price, quantity=qty
        )
        subtotal += product.price * qty
        if product.stock is not None:
            product.stock = max(0, product.stock - qty)
            product.save(update_fields=["stock"])

    order.total = subtotal * (1 - discount / 100)
    order.save(update_fields=["total"])
    return JsonResponse({"order_id": order.pk, "total": str(order.total)})


@login_required
@require_POST
def delete_order(request, pk):
    """Cancella un ordine e ripristina lo stock."""
    order = get_object_or_404(Order, pk=pk)
    for item in order.items.all():
        if item.product and item.product.stock is not None:
            item.product.stock += item.quantity
            item.product.save(update_fields=["stock"])
    order.delete()
    return JsonResponse({"ok": True})


@login_required
def print_order(request, pk):
    """Stampa l'ordine sulle stampanti configurate via ESC/POS."""
    order = get_object_or_404(Order, pk=pk)
    items = order.items.select_related("product").all()
    note = order.note

    # Stampante default (cassa) - stampa tutti i prodotti
    default_printer = Printer.objects.filter(is_default=True).first()

    # Raggruppa prodotti per stampante extra
    printer_items = {}  # printer_id -> [items]
    for item in items:
        if item.product:
            for printer in item.product.printers.all():
                printer_items.setdefault(printer.id, []).append(item)

    results = []

    # Stampa su cassa (default) - tutti i prodotti
    if default_printer:
        success, msg = _send_to_printer(default_printer, order, list(items), note, "Cassa")
        results.append({"printer": default_printer.name, "ok": success, "message": msg})

    # Stampa su stampanti extra - solo i prodotti assegnati
    for printer_id, p_items in printer_items.items():
        printer = Printer.objects.get(pk=printer_id)
        success, msg = _send_to_printer(printer, order, p_items, note, printer.name)
        results.append({"printer": printer.name, "ok": success, "message": msg})

    all_ok = all(r["ok"] for r in results)
    return JsonResponse({"ok": all_ok, "results": results}, status=200 if all_ok else 207)


def _send_to_printer(printer, order, items, note, dest_label):
    """Invia comanda ESC/POS a una stampante termica."""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(5)
        s.connect((printer.ip_address, printer.port))

        # ESC/POS commands
        ESC = b'\x1b'
        GS = b'\x1d'
        INIT = ESC + b'\x40'
        CENTER = ESC + b'\x61\x01'
        LEFT = ESC + b'\x61\x00'
        BOLD_ON = ESC + b'\x45\x01'
        BOLD_OFF = ESC + b'\x45\x00'
        DOUBLE_H = GS + b'\x21\x01'
        NORMAL = GS + b'\x21\x00'
        CUT = GS + b'\x56\x00'

        data = bytearray()
        data += INIT
        data += CENTER
        data += DOUBLE_H
        data += f"COMANDA #{order.pk}\n".encode()
        data += NORMAL
        data += f"{dest_label.upper()}\n".encode()
        from django.utils import timezone as tz
        data += f"{tz.localtime(order.created_at).strftime('%d/%m/%Y %H:%M')}\n".encode()
        data += b'\n'
        data += LEFT
        data += b'--------------------------------\n'
        data += BOLD_ON
        for item in items:
            data += f"{item.quantity}x {item.product_name}\n".encode()
        data += BOLD_OFF
        data += b'--------------------------------\n'
        if note:
            data += f"Note: {note}\n".encode()
            data += b'--------------------------------\n'
        data += b'\n\n\n'
        data += CUT

        s.sendall(bytes(data))
        s.close()
        return True, "OK"
    except Exception as e:
        return False, str(e)


# --- Inventario ---
@login_required
def inventario(request):
    products = Product.objects.select_related("category").prefetch_related("printers").order_by("name")
    categories = Category.objects.all()
    printers = Printer.objects.filter(is_default=False).order_by("name")
    return render(request, "cassa/inventario.html", {
        "products": products, "categories": categories,
        "icon_choices": Product.ICON_CHOICES, "printers": printers
    })


@login_required
@require_POST
def product_save(request):
    data = json.loads(request.body)
    pid = data.get("id")
    cat, _ = Category.objects.get_or_create(name=data["category"]) if data.get("category") else (None, False)

    price = Decimal(str(data["price"])) if data.get("price") else Decimal("0")
    stock = data.get("stock")
    if stock is not None and stock != "":
        stock = int(stock)
    else:
        stock = None

    if pid:
        p = get_object_or_404(Product, pk=pid)
        p.name = data["name"]
        p.price = price
        p.stock = stock
        p.is_shortcut = data.get("is_shortcut", p.is_shortcut)
        p.icon = data.get("icon", p.icon)
        p.print_destinations = data.get("print_destinations", p.print_destinations)
        p.category = cat
        p.save()
    else:
        p = Product.objects.create(
            name=data["name"], price=price, stock=stock,
            is_shortcut=data.get("is_shortcut", False), icon=data.get("icon", ""),
            print_destinations=data.get("print_destinations", ""), category=cat
        )
    # Aggiorna stampanti M2M
    printer_ids = data.get("printer_ids", [])
    p.printers.set(printer_ids)
    return JsonResponse({"id": p.pk})


@login_required
@require_POST
def product_delete(request, pk):
    get_object_or_404(Product, pk=pk).delete()
    return JsonResponse({"ok": True})


# --- Prelievo cassa ---
@login_required
@require_POST
def create_withdrawal(request):
    data = json.loads(request.body)
    operator = get_object_or_404(Operator, pk=data["operator_id"])
    amount = Decimal(str(data["amount"]))
    note = data.get("note", "")
    w = CashWithdrawal.objects.create(operator=operator, amount=amount, note=note)
    return JsonResponse({"id": w.pk, "operator": operator.name, "amount": str(w.amount)})


# --- Operatori ---
@login_required
def operatori(request):
    ops = Operator.objects.order_by("name")
    withdrawals = CashWithdrawal.objects.select_related("operator").order_by("-created_at")[:50]
    return render(request, "cassa/operatori.html", {"operators": ops, "withdrawals": withdrawals})


@login_required
@require_POST
def operator_save(request):
    data = json.loads(request.body)
    oid = data.get("id")
    if oid:
        op = get_object_or_404(Operator, pk=oid)
        op.name = data["name"]
        op.save()
    else:
        op = Operator.objects.create(name=data["name"])
    return JsonResponse({"id": op.pk})


@login_required
@require_POST
def operator_delete(request, pk):
    get_object_or_404(Operator, pk=pk).delete()
    return JsonResponse({"ok": True})


# --- Andamento ---
@login_required
def andamento(request):
    return render(request, "cassa/andamento.html")


# --- Ordini serata ---
@login_required
def ordini(request):
    today_session = get_session_date()
    start, end = session_range(today_session)
    orders = Order.objects.filter(created_at__gte=start, created_at__lt=end).prefetch_related("items").order_by("-created_at")
    return render(request, "cassa/ordini.html", {"orders": orders, "session_date": today_session})


@login_required
def andamento_data(request):
    days = int(request.GET.get("days", 7))
    start = timezone.now() - timedelta(days=days)
    orders = Order.objects.filter(created_at__gte=start)
    items = OrderItem.objects.filter(order__created_at__gte=start)

    daily = (
        orders.annotate(date=TruncDate("created_at"))
        .values("date")
        .annotate(total=Sum("total"), count=Count("id"))
        .order_by("date")
    )

    summary = orders.aggregate(total=Sum("total"), count=Count("id"))
    avg_order = float(summary["total"] or 0) / max(summary["count"], 1)

    by_category = (
        items.values(cat_name=F("product__category__name"))
        .annotate(qty=Sum("quantity"), revenue=Sum(F("price") * F("quantity")))
        .order_by("-revenue")
    )

    by_payment = (
        orders.values("payment_method")
        .annotate(total=Sum("total"), count=Count("id"))
        .order_by("-total")
    )

    top_products = (
        items.values("product_name")
        .annotate(qty=Sum("quantity"), revenue=Sum(F("price") * F("quantity")))
        .order_by("-qty")[:10]
    )

    return JsonResponse({
        "daily": [{"date": str(d["date"]), "total": str(d["total"] or 0), "count": d["count"]} for d in daily],
        "summary": {"total": str(summary["total"] or 0), "count": summary["count"], "avg": f"{avg_order:.2f}"},
        "by_category": [{"name": c["cat_name"] or "Senza categoria", "qty": c["qty"], "revenue": str(c["revenue"] or 0)} for c in by_category],
        "by_payment": [{"method": p["payment_method"], "total": str(p["total"] or 0), "count": p["count"]} for p in by_payment],
        "top_products": [{"name": t["product_name"], "qty": t["qty"], "revenue": str(t["revenue"] or 0)} for t in top_products],
    })


# --- Resoconti (Excel per serata) ---
@login_required
def resoconti(request):
    """Lista serate con ordini."""
    all_orders = Order.objects.all().order_by("-created_at")
    sessions = {}
    for order in all_orders:
        sd = get_session_date(timezone.localtime(order.created_at))
        if sd not in sessions:
            sessions[sd] = {"date": sd, "total": Decimal("0"), "count": 0}
        sessions[sd]["total"] += order.total
        sessions[sd]["count"] += 1

    days = sorted(sessions.values(), key=lambda x: x["date"], reverse=True)
    return render(request, "cassa/resoconti.html", {"days": days})


EURO_FMT = '#,##0.00 "\u20ac"'


def _add_category_sheet(wb, items_qs, sheet_title="Per Categoria"):
    ws = wb.create_sheet(title=sheet_title)
    ws.append(["Categoria", "Prodotto", "Qtà venduta", "Incasso"])
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 20

    by_cat_product = (
        items_qs.values(cat_name=F("product__category__name"), prod_name=F("product_name"))
        .annotate(qty=Sum("quantity"), revenue=Sum(F("price") * F("quantity")))
        .order_by("cat_name", "-qty")
    )
    for row in by_cat_product:
        ws.append([row["cat_name"] or "Senza categoria", row["prod_name"], row["qty"], round(float(row["revenue"] or 0), 2)])
        ws.cell(ws.max_row, 4).number_format = EURO_FMT

    ws.append([])
    ws.append(["RIEPILOGO PER CATEGORIA"])
    ws.append(["Categoria", "", "Qtà totale", "Incasso totale"])
    by_cat = (
        items_qs.values(cat_name=F("product__category__name"))
        .annotate(qty=Sum("quantity"), revenue=Sum(F("price") * F("quantity")))
        .order_by("-revenue")
    )
    for row in by_cat:
        ws.append([row["cat_name"] or "Senza categoria", "", row["qty"], round(float(row["revenue"] or 0), 2)])
        ws.cell(ws.max_row, 4).number_format = EURO_FMT


def _build_session_workbook(session_date):
    """Genera un workbook xlsx per una serata (05:00 -> 05:00 giorno dopo)."""
    start, end = session_range(session_date)
    orders = Order.objects.filter(created_at__gte=start, created_at__lt=end).prefetch_related("items")
    withdrawals = CashWithdrawal.objects.filter(created_at__gte=start, created_at__lt=end).select_related("operator")
    items = OrderItem.objects.filter(order__created_at__gte=start, order__created_at__lt=end)

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Ordini"
    headers = ["Ordine #", "Ora", "Prodotto", "Qtà", "Prezzo unit.", "Subtotale", "Pagamento", "Sconto %", "Totale ordine"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    for order in orders:
        for i, item in enumerate(order.items.all()):
            row = [
                order.pk,
                timezone.localtime(order.created_at).strftime("%H:%M"),
                item.product_name,
                item.quantity,
                round(float(item.price), 2),
                round(float(item.price * item.quantity), 2),
                order.get_payment_method_display() if i == 0 else "",
                float(order.discount_percent) if i == 0 else "",
                round(float(order.total), 2) if i == 0 else "",
            ]
            ws.append(row)
            r = ws.max_row
            for c in (5, 6):
                ws.cell(r, c).number_format = EURO_FMT
            if i == 0:
                ws.cell(r, 9).number_format = EURO_FMT

    if withdrawals.exists():
        ws.append([])
        ws.append(["PRELIEVI"])
        ws.append(["Ora", "Operatore", "Importo", "Note"])
        for w in withdrawals:
            ws.append([timezone.localtime(w.created_at).strftime("%H:%M"), w.operator.name, round(float(w.amount), 2), w.note])
            ws.cell(ws.max_row, 3).number_format = EURO_FMT

    ws.append([])
    total_orders = round(float(orders.aggregate(t=Sum("total"))["t"] or 0), 2)
    total_withdrawals = round(float(withdrawals.aggregate(t=Sum("amount"))["t"] or 0), 2)
    ws.append(["", "", "", "", "", "", "", "TOTALE ORDINI", total_orders])
    ws.cell(ws.max_row, 9).number_format = EURO_FMT
    ws.append(["", "", "", "", "", "", "", "TOTALE PRELIEVI", total_withdrawals])
    ws.cell(ws.max_row, 9).number_format = EURO_FMT
    ws.append(["", "", "", "", "", "", "", "NETTO", round(total_orders - total_withdrawals, 2)])
    ws.cell(ws.max_row, 9).number_format = EURO_FMT

    _add_category_sheet(wb, items)

    return wb


@login_required
def resoconti_download(request, date):
    session_date = datetime.strptime(date, "%Y-%m-%d").date()
    wb = _build_session_workbook(session_date)
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="resoconto_serata_{date}.xlsx"'
    wb.save(response)
    return response


@login_required
def resoconti_export_all(request):
    all_orders = Order.objects.all().order_by("created_at")
    session_dates = set()
    for order in all_orders:
        session_dates.add(get_session_date(timezone.localtime(order.created_at)))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sd in sorted(session_dates):
        start, end = session_range(sd)
        orders = Order.objects.filter(created_at__gte=start, created_at__lt=end).prefetch_related("items")
        ws = wb.create_sheet(title=str(sd))
        headers = ["Ordine #", "Ora", "Prodotto", "Qtà", "Prezzo unit.", "Subtotale", "Pagamento", "Sconto %", "Totale ordine"]
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14

        for order in orders:
            for i, item in enumerate(order.items.all()):
                row = [
                    order.pk,
                    timezone.localtime(order.created_at).strftime("%H:%M"),
                    item.product_name,
                    item.quantity,
                    round(float(item.price), 2),
                    round(float(item.price * item.quantity), 2),
                    order.get_payment_method_display() if i == 0 else "",
                    float(order.discount_percent) if i == 0 else "",
                    round(float(order.total), 2) if i == 0 else "",
                ]
                ws.append(row)
                r = ws.max_row
                for c in (5, 6):
                    ws.cell(r, c).number_format = EURO_FMT
                if i == 0:
                    ws.cell(r, 9).number_format = EURO_FMT

        ws.append([])
        total_val = round(float(orders.aggregate(t=Sum("total"))["t"] or 0), 2)
        ws.append(["", "", "", "", "", "", "", "TOTALE", total_val])
        ws.cell(ws.max_row, 9).number_format = EURO_FMT

    all_items = OrderItem.objects.all()
    _add_category_sheet(wb, all_items, "Riepilogo Categorie")

    if not wb.sheetnames:
        wb.create_sheet("Vuoto")

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="resoconto_completo.xlsx"'
    wb.save(response)
    return response


# --- Impostazioni ---
import socket


@login_required
def impostazioni(request):
    printers = Printer.objects.all().order_by("-is_default", "name")
    return render(request, "cassa/impostazioni.html", {"printers": printers})


@login_required
@require_POST
def printer_save(request):
    data = json.loads(request.body)
    pid = data.get("id")
    if data.get("is_default"):
        Printer.objects.filter(is_default=True).update(is_default=False)
    if pid:
        p = get_object_or_404(Printer, pk=pid)
        p.name = data["name"]
        p.ip_address = data["ip_address"]
        p.port = int(data.get("port", 9100))
        p.is_default = data.get("is_default", False)
        p.save()
    else:
        p = Printer.objects.create(
            name=data["name"], ip_address=data["ip_address"],
            port=int(data.get("port", 9100)), is_default=data.get("is_default", False)
        )
    return JsonResponse({"id": p.pk})


@login_required
@require_POST
def printer_delete(request, pk):
    get_object_or_404(Printer, pk=pk).delete()
    return JsonResponse({"ok": True})


@login_required
def printer_test(request, pk):
    printer = get_object_or_404(Printer, pk=pk)
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(3)
        s.connect((printer.ip_address, printer.port))
        # Invia test di stampa ESC/POS
        s.sendall(b"\x1b\x40")  # Init
        s.sendall(b"\x1b\x61\x01")  # Center
        s.sendall(b"=== TEST STAMPANTE ===\n")
        s.sendall(f"{printer.name}\n".encode())
        s.sendall(f"{printer.ip_address}:{printer.port}\n".encode())
        s.sendall(b"\n\n\n")
        s.sendall(b"\x1d\x56\x00")  # Cut
        s.close()
        return JsonResponse({"status": "ok", "message": f"Connessione riuscita a {printer.ip_address}:{printer.port}"})
    except Exception as e:
        return JsonResponse({"status": "error", "message": f"Errore: {str(e)}"}, status=400)


@login_required
def network_scan(request):
    """Scansiona la rete locale per tutti i dispositivi attivi."""
    import concurrent.futures
    import subprocess

    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        local_ip = s.getsockname()[0]
        s.close()
        subnet = ".".join(local_ip.split(".")[:3])
    except Exception:
        subnet = "192.168.1"
        local_ip = "sconosciuto"

    def check_host(i):
        ip = f"{subnet}.{i}"
        # Ping
        try:
            result = subprocess.run(["ping", "-c", "1", "-W", "1", ip], capture_output=True, timeout=2)
            if result.returncode != 0:
                return None
        except Exception:
            return None
        # Check porta 9100
        has_9100 = False
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            s.settimeout(0.5)
            has_9100 = s.connect_ex((ip, 9100)) == 0
            s.close()
        except Exception:
            pass
        return {"ip": ip, "printer": has_9100}

    found = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=50) as executor:
        futures = {executor.submit(check_host, i): i for i in range(1, 255)}
        for future in concurrent.futures.as_completed(futures):
            result = future.result()
            if result:
                found.append(result)

    found.sort(key=lambda x: [int(p) for p in x["ip"].split(".")])
    return JsonResponse({"devices": found, "subnet": subnet, "local_ip": local_ip})
